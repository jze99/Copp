from openpyxl import Workbook
from openpyxl.styles import Alignment

import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))


from src.data_temp import  data_employment_dynamics,data_FPM_dinamics



class create_xlsx_FPM_POO:
    def __init__(self):
        pass
        self.wb = Workbook()
        self.ws = self.wb.active
        self.temp = data_FPM_dinamics()
        self.temp.load_data()
        
    def add_data_POO(self):
        column_table = 1
        row_table = 1
        for idata ,data in enumerate(self.temp.model_data, start=1):
            self.ws.merge_cells(start_row=row_table, end_row=row_table, start_column=column_table, end_column=column_table+5)
            self.ws.cell(row=row_table,column=column_table).alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=row_table, column=column_table).value = data
            row_table+=1
            #sheet.append(["ПОО", "Сумма","проходят производственную практику","будут проходить производственную практику","трудоустроены на предприятие","Всего"])
            self.ws.cell(row=row_table, column=column_table).value = "ПОО"
            self.ws.cell(row=row_table, column=column_table+1).value = "Сумма"
            self.ws.cell(row=row_table, column=column_table+2).value = "проходят производственную практику"
            self.ws.cell(row=row_table, column=column_table+3).value = "будут проходить производственную практику"
            self.ws.cell(row=row_table, column=column_table+4).value = "трудоустроены на предприятие"
            self.ws.cell(row=row_table, column=column_table+5).value = "Всего"
            row_table+=1
            
            for imodel, model in enumerate(self.temp.model_data[data], start=3):
                temp_model = self.temp.model_data[data][model]
                self.ws.cell(row=row_table, column=column_table).value = temp_model.poo
                self.ws.cell(row=row_table, column=column_table+1).value = temp_model.summ
                self.ws.cell(row=row_table, column=column_table+2).value = temp_model.undergoing_production_practice
                self.ws.cell(row=row_table, column=column_table+3).value = temp_model.will_undergo_production_internship
                self.ws.cell(row=row_table, column=column_table+4).value = temp_model.employed_by_company
                self.ws.cell(row=row_table, column=column_table+5).value = temp_model.total
                row_table+=1
                pass
        
    def add_data_FPM(self):
        column_table=10
        row_table = 1
        
        self.ws.cell(row=row_table, column=column_table).value = "дата"
        self.ws.cell(row=row_table+1, column=column_table).value = "Сумма"
        self.ws.cell(row=row_table+2, column=column_table).value = "проходят производственную практику"
        self.ws.cell(row=row_table+3, column=column_table).value = "будут проходить производственную практику"
        self.ws.cell(row=row_table+4, column=column_table).value = "трудоустроены на предприятие"
        self.ws.cell(row=row_table+5, column=column_table).value = "Всего"
        
        column_table+=1
        
        for idata, data in enumerate(self.temp.poo_dinamics,start=1):
            for idinamic, dinamic in enumerate(self.temp.poo_dinamics[data], start=1):
                self.ws.cell(row=row_table, column=column_table).value = data
                self.ws.cell(row=row_table+1, column=column_table).value = dinamic.summ
                self.ws.cell(row=row_table+2, column=column_table).value = dinamic.undergoing_production_practice
                self.ws.cell(row=row_table+3, column=column_table).value = dinamic.will_undergo_production_internship
                self.ws.cell(row=row_table+4, column=column_table).value = dinamic.employed_by_company
                self.ws.cell(row=row_table+5, column=column_table).value = dinamic.total
                column_table+=1
        
    def create_file(self):
        self.add_data_POO()
        self.add_data_FPM()
        self.wb.save("example_FPM_POO.xlsx")



class create_xlsx_Employment:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.temp = data_employment_dynamics()
        self.temp.load_data()
    
    def add_data_spec_data(self):
        irow=1
        icolumn=1
        for data in self.temp.data_spec_data:
            self.ws.cell(row=irow, column=icolumn).value=data
            irow+=1
            self.ws.cell(row=irow, column=icolumn).value="код специальности"
            self.ws.cell(row=irow, column=icolumn+1).value="Суммарный выпуск"
            self.ws.cell(row=irow, column=icolumn+2).value="Трудоустроены"
            self.ws.cell(row=irow, column=icolumn+3).value="из них трудоустроены по полученной профессии, специальности"
            self.ws.cell(row=irow, column=icolumn+4).value="Зарегистрированы в качестве индивидуального предпринимателя или оформили самозанятость"
            self.ws.cell(row=irow, column=icolumn+5).value="Будут трудоустроены"
            self.ws.cell(row=irow, column=icolumn+6).value="из них будут трудоустроены по полученной профессии, специальности"
            self.ws.cell(row=irow, column=icolumn+7).value="Планируют зарегистрироваться в качестве индивидуального предпринимателя или оформить самозанятость"
            self.ws.cell(row=irow, column=icolumn+8).value="Продолжили (продолжат) обучение и не трудоустроились"
            self.ws.cell(row=irow, column=icolumn+9).value="Призваны (будут призваны) в Вооруженные Силы РФ"
            self.ws.cell(row=irow, column=icolumn+10).value="Находятся (будут находиться) в отпуске по уходу за ребенком"
            self.ws.cell(row=irow, column=icolumn+11).value="Находятся под следствием, отбывают наказание "
            self.ws.cell(row=irow, column=icolumn+12).value="Ухаживают за больными родственниками"
            self.ws.cell(row=irow, column=icolumn+13).value="Зарегистрированы в центрах занятости"
            self.ws.cell(row=irow, column=icolumn+14).value="Переехали"
            self.ws.cell(row=irow, column=icolumn+15).value="Не планируют трудоустраиваться"
            self.ws.cell(row=irow, column=icolumn+16).value="Тяжелое состояние здоровья, не позволяющее трудоустраиваться; смерть "
            self.ws.cell(row=irow, column=icolumn+17).value="Иные причины нахождения под риском нетрудоустройства"
            irow+=1
            for object in self.temp.data_spec_data[data]: 
                self.ws.cell(row=irow, column=icolumn).value=object
                self.ws.cell(row=irow, column=icolumn+1).value=self.temp.data_spec_data[data][object].The_total_output
                self.ws.cell(row=irow, column=icolumn+2).value=self.temp.data_spec_data[data][object].Employed
                self.ws.cell(row=irow, column=icolumn+3).value=self.temp.data_spec_data[data][object].of_these_employed_in_their_profession_specialty
                self.ws.cell(row=irow, column=icolumn+4).value=self.temp.data_spec_data[data][object].Registered_as_an_individual_entrepreneur_or_self_employed
                self.ws.cell(row=irow, column=icolumn+5).value=self.temp.data_spec_data[data][object].They_will_be_employed
                self.ws.cell(row=irow, column=icolumn+6).value=self.temp.data_spec_data[data][object].of_these_they_will_be_employed_in_their_profession_specialty
                self.ws.cell(row=irow, column=icolumn+7).value=self.temp.data_spec_data[data][object].They_plan_to_register_as_an_individual
                self.ws.cell(row=irow, column=icolumn+8).value=self.temp.data_spec_data[data][object].They_have_continued_their_studies_and_have_not_found_a_job
                self.ws.cell(row=irow, column=icolumn+9).value=self.temp.data_spec_data[data][object].Drafted_into_the_Armed_Forces_of_the_Russian_Federation
                self.ws.cell(row=irow, column=icolumn+10).value=self.temp.data_spec_data[data][object].Are_on_parental_leave
                self.ws.cell(row=irow, column=icolumn+11).value=self.temp.data_spec_data[data][object].They_are_under_investigation_and_are_serving_their_sentence
                self.ws.cell(row=irow, column=icolumn+12).value=self.temp.data_spec_data[data][object].They_take_care_of_sick_relatives
                self.ws.cell(row=irow, column=icolumn+13).value=self.temp.data_spec_data[data][object].Registered_in_employment_centers_as_unemployed
                self.ws.cell(row=irow, column=icolumn+14).value=self.temp.data_spec_data[data][object].Have_moved_outside_the_Russian_Federation
                self.ws.cell(row=irow, column=icolumn+15).value=self.temp.data_spec_data[data][object].They_do_not_plan_to_get_a_job_including
                self.ws.cell(row=irow, column=icolumn+16).value=self.temp.data_spec_data[data][object].Severe_health_condition_that_does_not_allow_employment_death
                self.ws.cell(row=irow, column=icolumn+17).value=self.temp.data_spec_data[data][object].Other_reasons_for_being_at_risk_of_disability
                irow+=1             
                
    def add_data_dinamic(self):
        irow=1
        icolumn=20
        self.ws.cell(row=irow+1, column=icolumn).value="Суммарный выпуск"
        self.ws.cell(row=irow+2, column=icolumn).value="Трудоустроены"
        self.ws.cell(row=irow+3, column=icolumn).value="из них трудоустроены по полученной профессии, специальности"
        self.ws.cell(row=irow+4, column=icolumn).value="Зарегистрированы в качестве индивидуального предпринимателя или оформили самозанятость"
        self.ws.cell(row=irow+5, column=icolumn).value="Будут трудоустроены"
        self.ws.cell(row=irow+6, column=icolumn).value="из них будут трудоустроены по полученной профессии, специальности"
        self.ws.cell(row=irow+7, column=icolumn).value="Планируют зарегистрироваться в качестве индивидуального предпринимателя или оформить самозанятость"
        self.ws.cell(row=irow+8, column=icolumn).value="Продолжили (продолжат) обучение и не трудоустроились"
        self.ws.cell(row=irow+9, column=icolumn).value="Призваны (будут призваны) в Вооруженные Силы РФ"
        self.ws.cell(row=irow+10, column=icolumn).value="Находятся (будут находиться) в отпуске по уходу за ребенком"
        self.ws.cell(row=irow+11, column=icolumn).value="Находятся под следствием, отбывают наказание "
        self.ws.cell(row=irow+12, column=icolumn).value="Ухаживают за больными родственниками"
        self.ws.cell(row=irow+13, column=icolumn).value="Зарегистрированы в центрах занятости"
        self.ws.cell(row=irow+14, column=icolumn).value="Переехали"
        self.ws.cell(row=irow+15, column=icolumn).value="Не планируют трудоустраиваться"
        self.ws.cell(row=irow+16, column=icolumn).value="Тяжелое состояние здоровья, не позволяющее трудоустраиваться; смерть "
        self.ws.cell(row=irow+17, column=icolumn).value="Иные причины нахождения под риском нетрудоустройства"
        icolumn+=1
        for data in self.temp.data_other_denamic:
            
            self.ws.cell(row=irow, column=icolumn).value=data
            
            self.ws.cell(row=irow+1, column=icolumn).value=self.temp.data_other_denamic[data].The_total_output
            self.ws.cell(row=irow+2, column=icolumn).value=self.temp.data_other_denamic[data].Employed
            self.ws.cell(row=irow+3, column=icolumn).value=self.temp.data_other_denamic[data].of_these_employed_in_their_profession_specialty
            self.ws.cell(row=irow+4, column=icolumn).value=self.temp.data_other_denamic[data].Registered_as_an_individual_entrepreneur_or_self_employed
            self.ws.cell(row=irow+5, column=icolumn).value=self.temp.data_other_denamic[data].They_will_be_employed
            self.ws.cell(row=irow+6, column=icolumn).value=self.temp.data_other_denamic[data].of_these_they_will_be_employed_in_their_profession_specialty
            self.ws.cell(row=irow+7, column=icolumn).value=self.temp.data_other_denamic[data].They_plan_to_register_as_an_individual
            self.ws.cell(row=irow+8, column=icolumn).value=self.temp.data_other_denamic[data].They_have_continued_their_studies_and_have_not_found_a_job
            self.ws.cell(row=irow+9, column=icolumn).value=self.temp.data_other_denamic[data].Drafted_into_the_Armed_Forces_of_the_Russian_Federation
            self.ws.cell(row=irow+10, column=icolumn).value=self.temp.data_other_denamic[data].Are_on_parental_leave
            self.ws.cell(row=irow+11, column=icolumn).value=self.temp.data_other_denamic[data].They_are_under_investigation_and_are_serving_their_sentence
            self.ws.cell(row=irow+12, column=icolumn).value=self.temp.data_other_denamic[data].They_take_care_of_sick_relatives
            self.ws.cell(row=irow+13, column=icolumn).value=self.temp.data_other_denamic[data].Registered_in_employment_centers_as_unemployed
            self.ws.cell(row=irow+14, column=icolumn).value=self.temp.data_other_denamic[data].Have_moved_outside_the_Russian_Federation
            self.ws.cell(row=irow+15, column=icolumn).value=self.temp.data_other_denamic[data].They_do_not_plan_to_get_a_job_including
            self.ws.cell(row=irow+16, column=icolumn).value=self.temp.data_other_denamic[data].Severe_health_condition_that_does_not_allow_employment_death
            self.ws.cell(row=irow+17, column=icolumn).value=self.temp.data_other_denamic[data].Other_reasons_for_being_at_risk_of_disability
            icolumn+=1
    
    def create_file(self):
        self.add_data_spec_data()
        self.add_data_dinamic()
        self.wb.save("example_employment.xlsx")
                
temp1 = create_xlsx_FPM_POO()
temp1.create_file()
temp2 = create_xlsx_Employment()
temp2.create_file()