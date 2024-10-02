from openpyxl import Workbook
from openpyxl.styles import Alignment
from data_temp import other_data

other_data.load_data()

class _create_xlsx:
    
    @classmethod
    def create_file(cls):
        wb = Workbook()
        sheet = wb.active

        column_table = 1
        row_table = 1
        for idata ,data in enumerate(other_data.model_data, start=1):
            sheet.merge_cells(start_row=row_table, end_row=row_table, start_column=column_table, end_column=column_table+5)
            sheet.cell(row=row_table,column=column_table).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=row_table, column=column_table).value = data
            row_table+=1
            #sheet.append(["ПОО", "Сумма","проходят производственную практику","будут проходить производственную практику","трудоустроены на предприятие","Всего"])
            sheet.cell(row=row_table, column=column_table).value = "ПОО"
            sheet.cell(row=row_table, column=column_table+1).value = "Сумма"
            sheet.cell(row=row_table, column=column_table+2).value = "проходят производственную практику"
            sheet.cell(row=row_table, column=column_table+3).value = "будут проходить производственную практику"
            sheet.cell(row=row_table, column=column_table+4).value = "трудоустроены на предприятие"
            sheet.cell(row=row_table, column=column_table+5).value = "Всего"
            row_table+=1
            
            for imodel, model in enumerate(other_data.model_data[data], start=3):
                temp_model = other_data.model_data[data][model]
                sheet.cell(row=row_table, column=column_table).value = temp_model.poo
                sheet.cell(row=row_table, column=column_table+1).value = temp_model.summ
                sheet.cell(row=row_table, column=column_table+2).value = temp_model.undergoing_production_practice
                sheet.cell(row=row_table, column=column_table+3).value = temp_model.will_undergo_production_internship
                sheet.cell(row=row_table, column=column_table+4).value = temp_model.employed_by_company
                sheet.cell(row=row_table, column=column_table+5).value = temp_model.total
                row_table+=1
                pass
            
        column_table+=8
        row_table = 1
        
        sheet.cell(row=row_table, column=column_table).value = "дата"
        sheet.cell(row=row_table+1, column=column_table).value = "Сумма"
        sheet.cell(row=row_table+2, column=column_table).value = "проходят производственную практику"
        sheet.cell(row=row_table+3, column=column_table).value = "будут проходить производственную практику"
        sheet.cell(row=row_table+4, column=column_table).value = "трудоустроены на предприятие"
        sheet.cell(row=row_table+5, column=column_table).value = "Всего"
        
        column_table+=1
        
        for idata, data in enumerate(other_data.poo_dinamics,start=1):
            for idinamic, dinamic in enumerate(other_data.poo_dinamics[data], start=1):
                sheet.cell(row=row_table, column=column_table).value = data
                sheet.cell(row=row_table+1, column=column_table).value = dinamic.summ
                sheet.cell(row=row_table+2, column=column_table).value = dinamic.undergoing_production_practice
                sheet.cell(row=row_table+3, column=column_table).value = dinamic.will_undergo_production_internship
                sheet.cell(row=row_table+4, column=column_table).value = dinamic.employed_by_company
                sheet.cell(row=row_table+5, column=column_table).value = dinamic.total
                column_table+=1
            
            
        wb.save("example.xlsx")

#create_xlsx.create_file()

class create_xlsx:
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def create_file_employment(self):
        
        for 