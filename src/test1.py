from openpyxl import Workbook
from openpyxl.styles import Font
from data_temp import load_data

temp = load_data()

wb = Workbook()
sheet = wb.active
for mass_data in temp:
    objects_list = list(mass_data.values())
    test=sheet.merge_cells(start_row=sheet.max_row+1,end_row=sheet.max_row+1,start_column=2, end_column=6)
    sheet.cell(row=sheet.max_row,column=2).value=objects_list[0].data
    sheet.append(["ПОО", "Сумма", "проходят производственную практику", "будут проходить производственную практику", "трудоустроены на предприятие", "Всего"])
    for idx, value in enumerate(objects_list, start=sheet.max_row+1):
        sheet.cell(row=idx, column=1).value = value.poo
        sheet.cell(row=idx, column=2).value = value.summ
        sheet.cell(row=idx, column=3).value = value.undergoing_production_practice
        sheet.cell(row=idx, column=4).value = value.will_undergo_production_internship
        sheet.cell(row=idx, column=5).value = value.employed_by_company
        sheet.cell(row=idx, column=6).value = value.total
wb.save("example.xlsx")
